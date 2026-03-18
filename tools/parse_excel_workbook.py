"""
parse_excel_workbook.py
-----------------------
Parses the CCA Attendance Predictor Excel workbook to extract:
  1. Tournament registration data (actual entry counts by date)
  2. Pre-fitted model parameters (from Model Summary sheet)
  3. Milestone predictions (from Prediction Timeline sheet)

Usage:
    python tools/parse_excel_workbook.py \\
        --file "CCA Attendance predictor working-IK-3 revised.xlsx" \\
        --tournament "World Open 2026" \\
        --out-data .tmp/cumulative_world_open.json \\
        --out-params .tmp/params_world_open.json

Output files:
    --out-data:   [[t, cumulative_count], ...]  (for fit_double_sigmoid.py)
    --out-params: {L1, k1, m1, L2, k2, m2, r2, rmse}  (pre-fitted from sheet)
"""

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── Sheet name heuristics ─────────────────────────────────────────────────────

SUMMARY_SHEET_HINTS = ["model summary", "summary", "parameters", "params"]
DATA_SHEET_HINTS    = ["actual data", "actual", "data", "residuals", "entries"]
TIMELINE_SHEET_HINTS = ["prediction timeline", "timeline", "predictions", "milestones"]


def find_sheet(wb, hints: list[str]):
    """Find a sheet whose name contains any of the hint strings (case-insensitive)."""
    names_lower = {name.lower(): name for name in wb.sheetnames}
    for hint in hints:
        for low_name, actual_name in names_lower.items():
            if hint in low_name:
                return wb[actual_name]
    return None


# ── Parameter extraction ──────────────────────────────────────────────────────

def extract_params(wb) -> dict | None:
    """
    Extract pre-fitted double-sigmoid parameters from the Model Summary sheet.
    Looks for cells labeled L1, k1, m1, L2, k2, m2, R², RMSE.
    """
    sheet = find_sheet(wb, SUMMARY_SHEET_HINTS)
    if sheet is None:
        print("Warning: No Model Summary sheet found", file=sys.stderr)
        return None

    params = {}
    param_map = {
        "l1": "L1", "k1": "k1", "m1": "m1",
        "l2": "L2", "k2": "k2", "m2": "m2",
        "r2": "r2", "r²": "r2", "r^2": "r2",
        "rmse": "rmse",
    }

    for row in sheet.iter_rows():
        for i, cell in enumerate(row):
            if cell.value is None:
                continue
            label = str(cell.value).strip().lower()
            if label in param_map:
                # Value is likely in the next cell to the right
                cells_in_row = list(row)
                if i + 1 < len(cells_in_row):
                    val_cell = cells_in_row[i + 1]
                    if isinstance(val_cell.value, (int, float)):
                        key = param_map[label]
                        params[key] = round(float(val_cell.value), 6)

    if len(params) < 6:
        print(f"Warning: Only found {len(params)} parameters in Model Summary sheet: {list(params.keys())}", file=sys.stderr)
        return params if params else None

    return params


# ── Registration data extraction ──────────────────────────────────────────────

def extract_registration_data(wb, first_reg_date: date | None = None) -> list[list]:
    """
    Extract [date, cumulative_count] or [t, cumulative_count] pairs from data sheet.

    Looks for columns named 'Date', 'Day', 't', 'Cumulative', 'Count', 'Entries', etc.
    Returns list of [t_value, cumulative_count].
    """
    sheet = find_sheet(wb, DATA_SHEET_HINTS)
    if sheet is None:
        print("Warning: No data sheet found", file=sys.stderr)
        return []

    # Find header row — first row with text content
    header_row_idx = None
    col_map = {}

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(isinstance(c, str) for c in row):
            header_row_idx = row_idx
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                label = str(cell).strip().lower()
                # Date / t column
                if any(k in label for k in ["date", " t ", "^t", "day", "days"]):
                    if "date" in label or "day" in label:
                        col_map["date"] = col_idx
                    else:
                        col_map["t"] = col_idx
                # Cumulative count column
                elif any(k in label for k in ["cumul", "total", "count", "entries", "registered"]):
                    if "cumul" in label or "total" in label:
                        col_map["cumulative"] = col_idx
                    else:
                        col_map.setdefault("count", col_idx)
            break

    if not col_map:
        print("Warning: Could not detect column headers in data sheet", file=sys.stderr)
        return []

    print(f"Detected columns: {col_map}", file=sys.stderr)

    data_points = []
    for row in sheet.iter_rows(min_row=(header_row_idx or 1) + 1, values_only=True):
        if all(c is None for c in row):
            continue

        t_val = None
        count_val = None

        # Resolve t-value
        if "t" in col_map:
            raw = row[col_map["t"]]
            if isinstance(raw, (int, float)):
                t_val = float(raw)
        elif "date" in col_map:
            raw = row[col_map["date"]]
            if isinstance(raw, datetime):
                reg_date = raw.date()
            elif isinstance(raw, date):
                reg_date = raw
            elif isinstance(raw, str):
                try:
                    reg_date = date.fromisoformat(raw)
                except ValueError:
                    continue
            else:
                continue

            if first_reg_date:
                t_val = (reg_date - first_reg_date).days
            else:
                # Return as date string; caller must convert
                t_val = reg_date.isoformat()

        # Resolve cumulative count
        cumul_key = "cumulative" if "cumulative" in col_map else "count"
        if cumul_key in col_map:
            raw = row[col_map[cumul_key]]
            if isinstance(raw, (int, float)):
                count_val = float(raw)

        if t_val is not None and count_val is not None and count_val > 0:
            data_points.append([t_val, count_val])

    return data_points


# ── Milestone extraction ──────────────────────────────────────────────────────

def extract_milestones(wb) -> list[dict]:
    """Extract milestone predictions from the Prediction Timeline sheet."""
    sheet = find_sheet(wb, TIMELINE_SHEET_HINTS)
    if sheet is None:
        return []

    milestones = []
    header_row_idx = None
    col_map = {}

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(isinstance(c, str) for c in row):
            header_row_idx = row_idx
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                label = str(cell).strip().lower()
                if "date" in label or "milestone" in label:
                    col_map["date"] = col_idx
                elif "label" in label or "event" in label or "description" in label:
                    col_map["label"] = col_idx
                elif "predicted" in label or "forecast" in label or "count" in label:
                    col_map["predicted"] = col_idx
            break

    if not col_map:
        return []

    for row in sheet.iter_rows(min_row=(header_row_idx or 1) + 1, values_only=True):
        if all(c is None for c in row):
            continue
        ms = {}
        if "date" in col_map and row[col_map["date"]]:
            raw = row[col_map["date"]]
            if isinstance(raw, datetime):
                ms["date"] = raw.date().isoformat()
            elif isinstance(raw, date):
                ms["date"] = raw.isoformat()
            elif isinstance(raw, str):
                ms["date"] = raw
        if "label" in col_map and row[col_map["label"]]:
            ms["label"] = str(row[col_map["label"]]).strip()
        if "predicted" in col_map and row[col_map["predicted"]]:
            raw = row[col_map["predicted"]]
            if isinstance(raw, (int, float)):
                ms["predicted"] = round(float(raw), 1)
        if ms.get("date"):
            milestones.append(ms)

    return milestones


# ── Entry list extraction (individual player rows) ────────────────────────────

def extract_entries(wb) -> list[dict]:
    """
    Look for a sheet with individual player entry rows.
    Returns [{name, registered_time}, ...] same format as scrape_entries.py output.
    """
    # Try all sheets if data sheet didn't work
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Look for "name" column
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if any(isinstance(c, str) and "name" in str(c).lower() for c in row):
                header = row
                name_col = next(
                    (i for i, c in enumerate(header) if c and "name" in str(c).lower()), None
                )
                time_col = next(
                    (i for i, c in enumerate(header)
                     if c and any(k in str(c).lower() for k in ["date", "time", "registered"])), None
                )
                if name_col is None:
                    continue

                entries = []
                for data_row in sheet.iter_rows(min_row=row_idx + 1, values_only=True):
                    if all(c is None for c in data_row):
                        continue
                    name = data_row[name_col] if name_col < len(data_row) else None
                    reg_time = None
                    if time_col is not None and time_col < len(data_row):
                        raw = data_row[time_col]
                        if isinstance(raw, datetime):
                            reg_time = raw.isoformat()
                        elif isinstance(raw, date):
                            reg_time = raw.isoformat()
                        elif isinstance(raw, str) and raw != "0000-00-00 00:00:00":
                            reg_time = raw
                    if name and str(name).strip():
                        entries.append({"name": str(name).strip(), "registered_time": reg_time})

                if len(entries) > 2:
                    print(f"Found {len(entries)} entries in sheet '{sheet_name}'", file=sys.stderr)
                    return entries

    return []


# ── CLI ───────────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(description="Parse CCA Excel workbook for tournament data")
    parser.add_argument("--file", required=True, help="Path to Excel workbook (.xlsx)")
    parser.add_argument("--tournament", default="Tournament", help="Tournament name for output")
    parser.add_argument("--first-reg-date", help="Date of first registration YYYY-MM-DD (for t-axis conversion)")
    parser.add_argument("--out-data", help="Output path for cumulative data JSON [[t, count], ...]")
    parser.add_argument("--out-params", help="Output path for model parameters JSON")
    parser.add_argument("--out-entries", help="Output path for individual entries JSON")
    parser.add_argument("--out-milestones", help="Output path for milestones JSON")
    parser.add_argument("--dump-sheets", action="store_true", help="Print all sheet names and exit")
    return parser.parse_args()


def main():
    args = parse_args()

    wb = openpyxl.load_workbook(args.file, data_only=True)
    print(f"Opened workbook: {args.file}", file=sys.stderr)
    print(f"Sheets: {wb.sheetnames}", file=sys.stderr)

    if args.dump_sheets:
        for name in wb.sheetnames:
            sheet = wb[name]
            print(f"\n=== {name} ({sheet.max_row} rows × {sheet.max_column} cols) ===")
            for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
                print(row)
        return

    first_reg = date.fromisoformat(args.first_reg_date) if args.first_reg_date else None

    # Extract all data
    params = extract_params(wb)
    data_points = extract_registration_data(wb, first_reg)
    entries = extract_entries(wb)
    milestones = extract_milestones(wb)

    # Save outputs
    def save(path, obj):
        if path:
            Path(path).parent.mkdir(parents=True, exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(obj, f, indent=2)
            print(f"Saved → {path}", file=sys.stderr)

    save(args.out_params, params)
    save(args.out_data, data_points)
    save(args.out_milestones, milestones)

    if args.out_entries:
        result = {
            "scraped_at": datetime.utcnow().isoformat() + "Z",
            "source": args.file,
            "tournament": args.tournament,
            "count": len(entries),
            "entries": entries,
        }
        save(args.out_entries, result)

    # Summary to stdout
    print(json.dumps({
        "tournament": args.tournament,
        "params_extracted": params is not None,
        "data_points": len(data_points),
        "entries": len(entries),
        "milestones": len(milestones),
    }, indent=2))


if __name__ == "__main__":
    main()
